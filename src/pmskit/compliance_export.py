# -*- coding: utf-8 -*-
"""Export a check_pms() result as a formatted ASME B31.3 schedule-adequacy report.

Columns: Item | Class | Size (NPS) | Schedule | Material | OD (mm) |
         Actual wall (mm) | Required (mm) | Margin (mm) | Status | Remark
"""
from __future__ import annotations
import re
from typing import Any, Dict, Optional

_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
COLUMNS = ["Item", "Class", "Size (NPS)", "Schedule", "Material", "OD (mm)",
           "Actual wall (mm)", "Required (mm)", "Margin (mm)", "Status", "Remark"]
_KEYS = ["item", "class", "size", "schedule", "material", "OD_mm",
         "actual_wall_mm", "required_mm", "margin_mm", "status", "remark"]


def _clean(v):
    return _ILLEGAL.sub(" ", v) if isinstance(v, str) else v


def to_compliance_xlsx(result: Dict[str, Any], path: str,
                       project: Optional[Dict[str, str]] = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = project or {}
    m, s = result["meta"], result["summary"]
    wb = Workbook(); ws = wb.active; ws.title = "B31.3 Schedule Check"

    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="1F3A5F"); hdr_font = Font(bold=True, color="FFFFFF")
    st_fill = {"UNDER-THICKNESS": PatternFill("solid", fgColor="FDE2E1"),
               "OK": PatternFill("solid", fgColor="E3F4E7"),
               "not-evaluated": PatternFill("solid", fgColor="EFF1F4"),
               "no-schedule": PatternFill("solid", fgColor="FFF4D6")}
    st_font = {"UNDER-THICKNESS": Font(bold=True, color="B91C1C"),
               "OK": Font(color="1F7A3D"), "not-evaluated": Font(color="6B7280"),
               "no-schedule": Font(color="92610A")}

    ws.merge_cells("A1:K1")
    ws["A1"] = project.get("title", "PIPING WALL-THICKNESS COMPLIANCE — ASME B31.3")
    ws["A1"].font = Font(bold=True, size=14); ws["A1"].alignment = Alignment(horizontal="center")

    lbl = Font(bold=True, color="374151")
    info = [
        ("Project:", project.get("project", "")),
        ("Company / PMS:", m.get("company", "")),
        ("Code:", m.get("code", "ASME B31.3")),
        ("Parameters:", f"E={m.get('E')}  W={m.get('W')}  mill tol={m.get('mill_tol')}"),
        ("Allowable stress source:", (s.get("materials_source") or "")
         + ("  [SYNTHETIC — replace with real datapack]" if s.get("synthetic_stress") else "")),
        ("Document No.:", project.get("doc_no", "")),
        ("Rev.:", project.get("rev", "")),
        ("Date:", project.get("date", "")),
        ("Summary:", f"{s['total']} PIPE checks — {s['ok']} OK, {s['under']} UNDER-THICKNESS, "
                     f"{s['not_evaluated']} not evaluated"),
    ]
    r = 3
    for k, v in info:
        ws.cell(r, 1, k).font = lbl
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=11)
        ws.cell(r, 2, _clean(v)); r += 1

    head = r + 1
    for j, name in enumerate(COLUMNS, 1):
        c = ws.cell(head, j, name); c.fill, c.font, c.border = hdr_fill, hdr_font, border
        c.alignment = Alignment(vertical="center", wrap_text=True)

    rr = head + 1
    for row in result["rows"]:
        for j, key in enumerate(_KEYS, 1):
            c = ws.cell(rr, j, _clean(row.get(key)))
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(key == "remark"))
        stc = ws.cell(rr, 10)
        st = row.get("status", "not-evaluated")
        stc.fill = st_fill.get(st, st_fill["not-evaluated"])
        stc.font = st_font.get(st, st_font["not-evaluated"])
        rr += 1

    for j, w in enumerate([6, 12, 10, 10, 20, 9, 13, 12, 11, 16, 44], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(head + 1, 1)
    ws.auto_filter.ref = f"A{head}:K{rr-1}"
    wb.save(path)
