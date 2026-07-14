# -*- coding: utf-8 -*-
"""Export a compare() result as a standard, consultant-style Deviation List.

Columns: Item | Class | Component | Size | Reference | Contractor |
         Deviation | Severity | Std Ref | Consultant Remark
"""
from __future__ import annotations
from typing import Any, Dict, Optional
import re

_ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
COLUMNS = ["Item", "Class", "Component", "Size", "Reference (baseline)",
           "Contractor", "Deviation", "Severity", "Std Ref", "Consultant Remark"]
_KEYS = ["item", "class", "component", "size", "reference", "contractor",
         "deviation", "severity", "std_ref", "remark"]


def _clean(v):
    return _ILLEGAL.sub(" ", str(v)) if isinstance(v, str) else v


def to_deviation_xlsx(result: Dict[str, Any], path: str,
                      project: Optional[Dict[str, str]] = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    project = project or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Deviation List"
    m, s = result["meta"], result["summary"]

    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=14)
    lbl = Font(bold=True, color="374151")
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    hdr_font = Font(bold=True, color="FFFFFF")
    sev_fill = {"major": PatternFill("solid", fgColor="FDE2E1"),
                "minor": PatternFill("solid", fgColor="FFF4D6"),
                "info": PatternFill("solid", fgColor="EAF1FB")}
    sev_font = {"major": Font(bold=True, color="B91C1C"),
                "minor": Font(color="92610A"), "info": Font(color="4B5563")}

    # --- title block ---
    ws.merge_cells("A1:J1")
    ws["A1"] = project.get("title", "PIPING MATERIAL SPECIFICATION — DEVIATION LIST")
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    info = [
        ("Project:", project.get("project", "")),
        ("Reference (baseline):", m.get("reference", "")),
        ("Contractor (detail design):", m.get("contractor", "")),
        ("Document No.:", project.get("doc_no", "")),
        ("Rev.:", project.get("rev", "")),
        ("Date:", project.get("date", "")),
        ("Baseline rule:", "Contractor is evaluated against the Reference PMS."),
        ("Summary:", f"{s['total']} deviations — {s['major']} major, {s['minor']} minor, "
                     f"{s['info']} info  |  Added {s['added']}, Removed {s['removed']}, Changed {s['changed']}"),
    ]
    r = 3
    for k, v in info:
        ws.cell(r, 1, k).font = lbl
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
        ws.cell(r, 2, _clean(v))
        r += 1

    # --- table header ---
    head_row = r + 1
    for j, name in enumerate(COLUMNS, start=1):
        c = ws.cell(head_row, j, name)
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
        c.alignment = Alignment(vertical="center", wrap_text=True)

    # --- rows ---
    rr = head_row + 1
    for row in result["rows"]:
        for j, key in enumerate(_KEYS, start=1):
            c = ws.cell(rr, j, _clean(row.get(key, "")))
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=(key in ("reference", "contractor", "remark")))
        sev = row.get("severity", "info")
        sc = ws.cell(rr, 8)
        sc.fill = sev_fill.get(sev, sev_fill["info"])
        sc.font = sev_font.get(sev, sev_font["info"])
        rr += 1

    # --- widths / freeze ---
    widths = [6, 12, 20, 10, 46, 46, 10, 9, 18, 40]
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(head_row + 1, 1)
    ws.auto_filter.ref = f"A{head_row}:J{rr-1}"
    wb.save(path)
